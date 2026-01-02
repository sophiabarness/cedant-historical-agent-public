"""Data loading utilities for CSV and Excel files."""

import csv
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple


def detect_csv_delimiter(file_path: Path, sample_size: int = 1024) -> str:
    """
    Detect the delimiter used in a CSV file.
    
    Args:
        file_path: Path to CSV file
        sample_size: Number of bytes to sample
        
    Returns:
        Detected delimiter character (defaults to comma)
    """
    try:
        with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
            sample = csvfile.read(sample_size)
            
            if not sample.strip():
                return ','
            
            sniffer = csv.Sniffer()
            delimiter = sniffer.sniff(sample).delimiter
            return delimiter
    except (csv.Error, Exception):
        return ','  # Default to comma if detection fails


def load_csv_file(
    file_path: Path,
    encoding: str = 'utf-8'
) -> Tuple[Optional[csv.DictReader], Optional[str]]:
    """
    Load a CSV file and return a DictReader.
    
    Args:
        file_path: Path to CSV file
        encoding: File encoding
        
    Returns:
        Tuple of (DictReader, error_message)
        Returns (None, error_message) if loading fails
    """
    try:
        # Validate file exists and is readable
        if not file_path.exists():
            return None, f"File not found: {file_path}"
        
        # Test file is readable
        try:
            with open(file_path, 'r', newline='', encoding=encoding) as test_file:
                test_file.read(100)
        except UnicodeDecodeError:
            return None, f"File has invalid encoding: {file_path}"
        except PermissionError:
            return None, f"Permission denied: {file_path}"
        
        # Detect delimiter
        delimiter = detect_csv_delimiter(file_path)
        
        # Open file and create reader
        csvfile = open(file_path, 'r', newline='', encoding=encoding)
        reader = csv.DictReader(csvfile, delimiter=delimiter)
        
        # Validate has headers
        if not reader.fieldnames:
            csvfile.close()
            return None, "CSV file has no header row"
        
        return reader, None
        
    except Exception as e:
        return None, f"Error loading CSV file: {str(e)}"


def load_excel_file(
    file_path: Path,
    sheet_name: Optional[str] = None,
    read_only: bool = True
):
    """
    Load an Excel file and return workbook and worksheet.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Optional specific sheet name (uses first sheet if None)
        read_only: Whether to open in read-only mode
        
    Returns:
        Tuple of (workbook, worksheet, error_message)
        Returns (None, None, error_message) if loading fails
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError:
        return None, None, "openpyxl library not available"
    
    try:
        # Validate file exists
        if not file_path.exists():
            return None, None, f"File not found: {file_path}"
        
        # Load workbook
        try:
            workbook = load_workbook(file_path, read_only=read_only, data_only=True)
        except InvalidFileException:
            return None, None, f"Invalid or corrupted Excel file: {file_path}"
        except PermissionError:
            return None, None, f"Permission denied: {file_path}"
        
        # Validate has sheets
        if not workbook.sheetnames:
            workbook.close()
            return None, None, "Excel file contains no worksheets"
        
        # Get worksheet
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                workbook.close()
                return None, None, f"Sheet '{sheet_name}' not found in workbook"
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook[workbook.sheetnames[0]]
        
        # Validate sheet has data
        if not worksheet.max_row or worksheet.max_row < 2:
            workbook.close()
            return None, None, f"Sheet '{worksheet.title}' appears to be empty"
        
        return workbook, worksheet, None
        
    except Exception as e:
        return None, None, f"Error loading Excel file: {str(e)}"


def find_data_sheet(workbook, preferred_keywords: List[str] = None) -> Optional[str]:
    """
    Find the most likely data sheet in a workbook.
    
    Args:
        workbook: Excel workbook object
        preferred_keywords: Optional list of keywords to prioritize
        
    Returns:
        Sheet name or None
    """
    if not preferred_keywords:
        preferred_keywords = ['data', 'lookup', 'table', 'loss', 'program', 'event', 'historical']
    
    # Look for sheets with preferred keywords
    for sheet_name in workbook.sheetnames:
        sheet_lower = sheet_name.lower()
        if any(keyword in sheet_lower for keyword in preferred_keywords):
            return sheet_name
    
    # Default to first sheet
    return workbook.sheetnames[0] if workbook.sheetnames else None


def validate_file_format(file_path: Path, allowed_extensions: List[str]) -> Tuple[bool, Optional[str]]:
    """
    Validate that a file has an allowed extension.
    
    Args:
        file_path: Path to file
        allowed_extensions: List of allowed extensions (e.g., ['.csv', '.xlsx'])
        
    Returns:
        Tuple of (is_valid, error_message)
    """
    file_extension = file_path.suffix.lower()
    
    if file_extension not in allowed_extensions:
        return False, f"Unsupported file format: {file_extension}. Supported: {', '.join(allowed_extensions)}"
    
    return True, None
