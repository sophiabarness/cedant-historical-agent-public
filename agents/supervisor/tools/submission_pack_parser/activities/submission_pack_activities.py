"""Activities for submission pack data extraction."""

import re
from temporalio import activity

from models.submission_pack import AsOfYearInput, AsOfYearOutput
from shared.config import get_submission_packs_dir


async def _signal_extraction_data(
    bridge_workflow_id: str,
    data_type: str,
    value: any
) -> bool:
    """
    Signal extraction data to the BridgeWorkflow.
    
    Args:
        bridge_workflow_id: The workflow ID of the BridgeWorkflow
        data_type: Type of data ("as_of_year" or "events")
        value: The actual data to store
        
    Returns:
        bool: True if signal was sent successfully, False otherwise
    """
    try:
        from shared.config import get_temporal_client
        
        client = await get_temporal_client()
        workflow_handle = client.get_workflow_handle(bridge_workflow_id)
        
        await workflow_handle.signal(
            "store_extraction_data",
            {
                "type": data_type,
                "value": value
            }
        )
        
        activity.logger.info(f"Signaled {data_type}='{value}' to BridgeWorkflow {bridge_workflow_id}")
        return True
        
    except Exception as e:
        activity.logger.warning(f"Failed to signal {data_type} to BridgeWorkflow: {str(e)}")
        return False


@activity.defn
async def extract_as_of_year(input_data: AsOfYearInput) -> AsOfYearOutput:
    """
    Temporal Activity for extracting As Of Year from Excel submission pack documents.
    
    This activity implements pattern matching for "As Of:", "Effective As Of:", 
    and date patterns, with worksheet scanning for table of contents and cover pages.
    Provides confidence scoring and source location tracking.
    
    Args:
        input_data: AsOfYearInput containing file path and optional bridge_workflow_id
        
    Returns:
        AsOfYearOutput with extraction results including confidence and source location
    """
    activity.logger.info(f"Extracting As Of Year from {input_data.file_path}")
    
    try:
        # Auto-detect file type from extension
        from pathlib import Path
        file_ext = Path(input_data.file_path).suffix.lower()
        
        if file_ext in ['.xlsx', '.xls', '.xlsm']:
            result = await _extract_as_of_year_excel(input_data.file_path)
        else:
            return AsOfYearOutput(
                success=False,
                error_message=f"Unsupported file extension: {file_ext}. Only Excel files (.xlsx, .xls, .xlsm) are supported."
            )
        
        # Signal to BridgeWorkflow if successful and workflow ID provided
        if result.success and result.as_of_year and input_data.bridge_workflow_id:
            await _signal_extraction_data(
                bridge_workflow_id=input_data.bridge_workflow_id,
                data_type="as_of_year",
                value=result.as_of_year
            )
        
        return result
        
    except Exception as e:
        activity.logger.error(f"Error extracting As Of Year: {str(e)}")
        return AsOfYearOutput(
            success=False,
            error_message=f"Error extracting As Of Year: {str(e)}"
        )


async def _extract_as_of_year_excel(file_path: str) -> AsOfYearOutput:
    """Extract As Of Year from Excel file with comprehensive pattern matching."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return AsOfYearOutput(
            success=False,
            error_message="openpyxl library not available for Excel processing"
        )
    
    try:
        workbook = load_workbook(file_path, read_only=True)
        activity.logger.info(f"Loaded Excel workbook with {len(workbook.sheetnames)} sheets")
        
        # Enhanced patterns for As Of Year detection
        as_of_patterns = [
            # Primary patterns with high confidence
            r'as\s+of[:\s]+.*?(\d{4})',
            r'effective\s+as\s+of[:\s]+.*?(\d{4})',
            r'as\s+of\s+date[:\s]+.*?(\d{4})',
            r'effective\s+date[:\s]+.*?(\d{4})',
            # Secondary patterns with medium confidence
            r'effective[:\s]+.*?(\d{4})',
            r'(\d{4})\s+as\s+of',
            r'renewal\s+date[:\s]+.*?(\d{4})',
            r'policy\s+year[:\s]+.*?(\d{4})',
            # Date format patterns
            r'(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})',
            r'(\d{4})[\/\-](\d{1,2})[\/\-](\d{1,2})',
        ]
        
        # Prioritize sheets likely to contain As Of Year information
        priority_sheets = []
        regular_sheets = []
        
        for sheet_name in workbook.sheetnames:
            sheet_lower = sheet_name.lower()
            if any(keyword in sheet_lower for keyword in ['toc', 'contents', 'cover', 'summary', 'index', 'overview']):
                priority_sheets.append((sheet_name, "high"))
            elif any(keyword in sheet_lower for keyword in ['info', 'general', 'details', 'submission']):
                priority_sheets.append((sheet_name, "medium"))
            else:
                regular_sheets.append((sheet_name, "low"))
        
        all_sheets = priority_sheets + regular_sheets[:2]
        
        best_match = None
        best_confidence = "low"
        
        for sheet_name, sheet_priority in all_sheets:
            activity.logger.info(f"Searching sheet: {sheet_name} (priority: {sheet_priority})")
            worksheet = workbook[sheet_name]
            
            search_rows = min(25, worksheet.max_row) if worksheet.max_row else 25
            search_cols = min(15, worksheet.max_column) if worksheet.max_column else 15
            
            for row in range(1, search_rows + 1):
                for col in range(1, search_cols + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        cell_text = str(cell.value).strip()
                        cell_lower = cell_text.lower()
                        
                        if len(cell_text) < 4:
                            continue
                        
                        for pattern_idx, pattern in enumerate(as_of_patterns):
                            match = re.search(pattern, cell_lower, re.IGNORECASE)
                            if match:
                                year = None
                                if pattern_idx < 8:
                                    year = match.group(1)
                                else:
                                    groups = match.groups()
                                    for group in groups:
                                        if len(group) == 4 and group.isdigit():
                                            year = group
                                            break
                                
                                if year and year.isdigit():
                                    year_int = int(year)
                                    if 2015 <= year_int <= 2030:
                                        confidence = determine_confidence_level(
                                            pattern_idx, sheet_priority, cell_text, row, col
                                        )
                                        
                                        source_location = f"Sheet: {sheet_name}, Cell: {cell.coordinate}"
                                        
                                        if confidence == "high":
                                            activity.logger.info(f"High confidence As Of Year found: {year} at {source_location}")
                                            workbook.close()
                                            return AsOfYearOutput(
                                                success=True,
                                                as_of_year=year,
                                                source_location=source_location,
                                                confidence_level=confidence,
                                                extracted_text=cell_text
                                            )
                                        
                                        if not best_match or is_better_match(confidence, best_confidence):
                                            best_match = {
                                                "year": year,
                                                "source_location": source_location,
                                                "confidence": confidence,
                                                "extracted_text": cell_text
                                            }
                                            best_confidence = confidence
        
        workbook.close()
        
        if best_match:
            activity.logger.info(f"Best As Of Year match found: {best_match['year']} with confidence: {best_match['confidence']}")
            return AsOfYearOutput(
                success=True,
                as_of_year=best_match["year"],
                source_location=best_match["source_location"],
                confidence_level=best_match["confidence"],
                extracted_text=best_match["extracted_text"]
            )
        
        activity.logger.warning("As Of Year not found in Excel document")
        return AsOfYearOutput(
            success=False,
            error_message="As Of Year not found in document"
        )
        
    except Exception as e:
        activity.logger.error(f"Error processing Excel file: {str(e)}")
        return AsOfYearOutput(
            success=False,
            error_message=f"Error processing Excel file: {str(e)}"
        )



# Confidence level functions are defined at the end of this file


# Duplicate activity definition removed - using the working implementation below

@activity.defn
async def locate_submission_pack_activity(input_data: dict) -> dict:
    """
    Temporal Activity for locating submission pack files.
    
    This activity handles file system operations to locate submission pack files
    based on Program ID, avoiding workflow restrictions on path operations.
    
    Args:
        input_data: Dict containing program_id and submission_packs_directory
        
    Returns:
        Dict containing file location results
    """
    activity.logger.info(f"Locating submission pack for Program ID: {input_data.get('program_id')}")
    
    try:
        from agents.supervisor.tools.submission_pack_parser.tools.locate_submission_pack import locate_submission_pack
        from models.submission_pack import FileLocatorInput
        
        program_id = input_data.get("program_id")
        submission_packs_directory = input_data.get("submission_packs_directory") or get_submission_packs_dir()
        
        # Create input model
        locator_input = FileLocatorInput(
            program_id=program_id,
            submission_packs_directory=submission_packs_directory
        )
        
        # Execute file location
        result = locate_submission_pack(locator_input)
        
        # Convert to dict format
        if result.success:
            activity.logger.info(f"Found submission pack: {result.file_name} at {result.file_path}")
            return {
                "success": True,
                "file_path": result.file_path,
                "file_name": result.file_name,
                "file_type": result.file_type,
                "program_id": program_id
            }
        else:
            activity.logger.warning(f"Submission pack not found: {result.error_message}")
            return {
                "success": False,
                "error_message": result.error_message,
                "program_id": program_id
            }
        
    except Exception as e:
        activity.logger.error(f"Error in locate submission pack activity: {str(e)}")
        return {
            "success": False,
            "error_message": f"File location failed: {str(e)}",
            "program_id": input_data.get("program_id", "unknown")
        }

@activity.defn
async def llm_extract_catastrophe_data_activity(input_data: dict) -> dict:
    """
    Temporal Activity for LLM-based catastrophe data extraction with calculated values.
    
    Args:
        input_data: Dictionary containing:
            - file_path: Path to submission pack file
            - sheet_names: List of sheet names to extract from
            - extraction_approach: Optional extraction approach description
            - user_instructions: Optional user-provided instructions to customize extraction
            - bridge_workflow_id: Optional workflow ID for signaling extracted data to BridgeWorkflow
    
    Returns:
        Dictionary with extraction results including calculated loss amounts
    """
    try:
        # Import inside activity to avoid non-deterministic imports in workflow context
        from agents.supervisor.tools.submission_pack_parser.tools.llm_extractor import LLMPrimaryExtractor
        from dataclasses import asdict
        
        file_path = input_data.get("file_path")
        sheet_names = input_data.get("sheet_names", [])
        extraction_approach = input_data.get("extraction_approach", "LLM-guided extraction with calculated values")
        user_instructions = input_data.get("user_instructions")
        bridge_workflow_id = input_data.get("bridge_workflow_id")
        
        activity.logger.info(f"Starting extraction: file_path={file_path}, sheet_names={sheet_names}")
        
        if not file_path or not sheet_names:
            activity.logger.error("Missing required parameters")
            return {
                "success": False,
                "error": "Missing required parameters: file_path and sheet_names"
            }
        
        # Call the extractor
        extractor = LLMPrimaryExtractor()
        if user_instructions:
            activity.logger.info(f"User instructions: {user_instructions}")
        
        result = extractor.extract_catastrophe_data(
            file_path=file_path,
            sheet_names=sheet_names,
            extraction_approach=extraction_approach,
            user_instructions=user_instructions
        )
        
        activity.logger.info(f"Extractor result: success={result.success}, events={len(result.events)}, error={result.error_message}")
        
        # Convert to dictionary format expected by the tool interface
        converted_events = [asdict(event) for event in result.events]
        activity.logger.info(f"Converted {len(converted_events)} events to dict format")
        
        # Signal events to BridgeWorkflow if successful and workflow ID provided
        if result.success and converted_events and bridge_workflow_id:
            try:
                signal_success = await _signal_extraction_data(
                    bridge_workflow_id=bridge_workflow_id,
                    data_type="events",
                    value=converted_events
                )
                if signal_success:
                    activity.logger.info(f"Successfully signaled {len(converted_events)} events to BridgeWorkflow")
                else:
                    activity.logger.warning("Signal to BridgeWorkflow returned False, continuing with extraction result")
            except Exception as signal_error:
                activity.logger.warning(f"Failed to signal events to BridgeWorkflow: {str(signal_error)}")
        
        return {
            "success": result.success,
            "events": converted_events,
            "extracted_count": len(result.events),
            "extraction_approach": result.extraction_approach,
            "notes": result.notes,
            "error_message": result.error_message
        }
        
    except Exception as e:
        activity.logger.error(f"Exception: {str(e)}")
        import traceback
        activity.logger.error(f"Traceback: {traceback.format_exc()}")
        return {
            "success": False,
            "events": [],
            "error": f"LLM extraction failed: {str(e)}"
        }

# Duplicate activity definition removed - using the working implementation above


def determine_confidence_level(pattern_idx: int, sheet_priority: str, cell_text: str, row: int, col: int) -> str:
    """
    Determine confidence level for As Of Year extraction based on multiple factors.
    
    Args:
        pattern_idx: Index of the pattern that matched (lower = higher confidence)
        sheet_priority: Priority level of the sheet ("high", "medium", "low")
        cell_text: The actual cell text that was matched
        row: Row number of the match
        col: Column number of the match
        
    Returns:
        str: Confidence level ("high", "medium", "low")
    """
    confidence_score = 0
    
    # Pattern-based scoring (primary patterns get higher scores)
    if pattern_idx < 4:  # Primary patterns like "as of", "effective as of"
        confidence_score += 3
    elif pattern_idx < 8:  # Secondary text patterns
        confidence_score += 2
    else:  # Date format patterns
        confidence_score += 1
    
    # Sheet priority scoring
    if sheet_priority == "high":
        confidence_score += 2
    elif sheet_priority == "medium":
        confidence_score += 1
    
    # Position scoring (earlier rows/columns more likely to contain As Of Year)
    if row <= 5:
        confidence_score += 2
    elif row <= 10:
        confidence_score += 1
    
    if col <= 3:
        confidence_score += 1
    
    # Text content scoring
    cell_lower = cell_text.lower()
    if any(keyword in cell_lower for keyword in ['as of', 'effective', 'renewal', 'policy']):
        confidence_score += 2
    
    # Convert score to confidence level
    if confidence_score >= 6:
        return "high"
    elif confidence_score >= 3:
        return "medium"
    else:
        return "low"


def is_better_match(new_confidence: str, current_confidence: str) -> bool:
    """
    Compare confidence levels to determine if new match is better.
    
    Args:
        new_confidence: Confidence level of new match
        current_confidence: Confidence level of current best match
        
    Returns:
        bool: True if new match is better
    """
    confidence_order = {"high": 3, "medium": 2, "low": 1}
    return confidence_order.get(new_confidence, 0) > confidence_order.get(current_confidence, 0)