"""
Sheet identification tools for intelligent sheet analysis.

These tools provide the core functionality for the sheet identification agent
to analyze Excel workbooks and identify catastrophe loss data sheets.
"""

from typing import Dict, Any, List, Optional
from pathlib import Path
from dataclasses import dataclass
from models.core import ToolDefinition, ToolArgument

try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


@dataclass
class SheetStructureResult:
    """Result from sheet structure analysis."""
    success: bool
    sheet_names: List[str]
    total_sheets: int
    file_path: str
    file_size_mb: Optional[float] = None
    error_message: str = ""


@dataclass
class SheetContentResult:
    """Result from reading sheet content."""
    success: bool
    sheet_name: str
    headers: List[str]
    data_rows: List[List[str]]
    total_rows: int
    total_columns: int
    filtered_columns: int = 0
    filtered_rows: int = 0
    error_message: str = ""


def get_sheet_names_tool(file_path: str) -> Dict[str, Any]:
    """
    Extract all sheet names and basic workbook metadata from an Excel file.
    
    This tool provides the sheet identification agent with a complete overview
    of the workbook structure, including all sheet names and basic metadata.
    
    Args:
        file_path: Path to the Excel file to analyze
        
    Returns:
        Dict containing:
            - success (bool): Whether the operation succeeded
            - sheet_names (List[str]): List of all sheet names in the workbook
            - total_sheets (int): Total number of sheets
            - file_path (str): Path to the analyzed file
            - file_size_mb (float): File size in megabytes
            - error_message (str): Error message if operation failed
    """
    try:
        if not EXCEL_AVAILABLE:
            return {
                "success": False,
                "sheet_names": [],
                "total_sheets": 0,
                "file_path": file_path,
                "file_size_mb": None,
                "error_message": "openpyxl library not available for Excel processing"
            }
        
        # Validate file exists
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            return {
                "success": False,
                "sheet_names": [],
                "total_sheets": 0,
                "file_path": file_path,
                "file_size_mb": None,
                "error_message": f"File not found: {file_path}"
            }
        
        # Get file size
        file_size_mb = file_path_obj.stat().st_size / (1024 * 1024)
        
        # Load workbook to get sheet names
        try:
            workbook = load_workbook(file_path, read_only=True)
        except InvalidFileException:
            return {
                "success": False,
                "sheet_names": [],
                "total_sheets": 0,
                "file_path": file_path,
                "file_size_mb": file_size_mb,
                "error_message": f"Invalid or corrupted Excel file: {file_path}"
            }
        except PermissionError:
            return {
                "success": False,
                "sheet_names": [],
                "total_sheets": 0,
                "file_path": file_path,
                "file_size_mb": file_size_mb,
                "error_message": f"Permission denied accessing file: {file_path}"
            }
        
        # Extract sheet information
        sheet_names = workbook.sheetnames
        total_sheets = len(sheet_names)
        
        # Close workbook to free resources
        workbook.close()
        
        return {
            "success": True,
            "sheet_names": sheet_names,
            "total_sheets": total_sheets,
            "file_path": file_path,
            "file_size_mb": round(file_size_mb, 2),
            "error_message": ""
        }
        
    except Exception as e:
        return {
            "success": False,
            "sheet_names": [],
            "total_sheets": 0,
            "file_path": file_path,
            "file_size_mb": None,
            "error_message": f"Unexpected error analyzing workbook structure: {str(e)}"
        }


def read_sheet_tool(file_path: str, sheet_name: str, mode: str = "preview") -> Dict[str, Any]:
    """
    Read content from a specific sheet in an Excel file with intelligent filtering.
    
    This tool allows the sheet identification agent to examine the actual content
    of sheets to make informed decisions about which sheets contain catastrophe data.
    Implements advanced filtering to remove empty columns and rows, providing
    clean, LLM-friendly data presentation.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read
        mode: Reading mode - "preview" (first 15 rows) or "full" (all rows)
        
    Returns:
        Dict containing:
            - success (bool): Whether the operation succeeded
            - sheet_name (str): Name of the sheet that was read
            - headers (List[str]): Column headers from the first row (filtered)
            - data_rows (List[List[str]]): Sample data rows (cleaned and filtered)
            - total_rows (int): Total number of rows in the sheet
            - total_columns (int): Total number of columns in the sheet
            - filtered_columns (int): Number of columns after filtering empty ones
            - filtered_rows (int): Number of rows after filtering empty ones
            - error_message (str): Error message if operation failed
    """
    try:
        if not EXCEL_AVAILABLE:
            return {
                "success": False,
                "sheet_name": sheet_name,
                "headers": [],
                "data_rows": [],
                "total_rows": 0,
                "total_columns": 0,
                "filtered_columns": 0,
                "filtered_rows": 0,
                "error_message": "openpyxl library not available for Excel processing"
            }
        
        # Validate file exists
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            return {
                "success": False,
                "sheet_name": sheet_name,
                "headers": [],
                "data_rows": [],
                "total_rows": 0,
                "total_columns": 0,
                "filtered_columns": 0,
                "filtered_rows": 0,
                "error_message": f"File not found: {file_path}"
            }
        
        # Load workbook
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except InvalidFileException:
            return {
                "success": False,
                "sheet_name": sheet_name,
                "headers": [],
                "data_rows": [],
                "total_rows": 0,
                "total_columns": 0,
                "filtered_columns": 0,
                "filtered_rows": 0,
                "error_message": f"Invalid or corrupted Excel file: {file_path}"
            }
        except PermissionError:
            return {
                "success": False,
                "sheet_name": sheet_name,
                "headers": [],
                "data_rows": [],
                "total_rows": 0,
                "total_columns": 0,
                "filtered_columns": 0,
                "filtered_rows": 0,
                "error_message": f"Permission denied accessing file: {file_path}"
            }
        
        # Check if sheet exists
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            return {
                "success": False,
                "sheet_name": sheet_name,
                "headers": [],
                "data_rows": [],
                "total_rows": 0,
                "total_columns": 0,
                "filtered_columns": 0,
                "filtered_rows": 0,
                "error_message": f"Sheet '{sheet_name}' not found in workbook. Available sheets: {', '.join(workbook.sheetnames)}"
            }
        
        # Get the worksheet
        worksheet = workbook[sheet_name]
        
        # Get sheet dimensions
        total_rows = worksheet.max_row or 0
        total_columns = worksheet.max_column or 0
        
        # Determine how many rows to read
        if mode == "full":
            max_rows_to_read = total_rows
        else:  # preview mode
            max_rows_to_read = min(15, total_rows)
        
        # First pass: Read all data to identify empty columns and rows
        raw_data = []
        for row_num in range(1, min(max_rows_to_read + 1, total_rows + 1)):
            row_data = []
            for col in range(1, min(total_columns + 1, 50)):  # Limit to 50 columns max
                cell_value = worksheet.cell(row=row_num, column=col).value
                cell_str = str(cell_value).strip() if cell_value is not None else ""
                row_data.append(cell_str)
            raw_data.append(row_data)
        
        # Identify completely empty columns
        non_empty_columns = []
        if raw_data:
            for col_idx in range(len(raw_data[0])):
                has_content = False
                for row in raw_data:
                    if col_idx < len(row) and row[col_idx] and row[col_idx].strip():
                        has_content = True
                        break
                if has_content:
                    non_empty_columns.append(col_idx)
        
        # Filter out empty columns from all rows and remove empty strings
        filtered_data = []
        for row in raw_data:
            filtered_row = []
            for col_idx in non_empty_columns:
                cell_value = row[col_idx] if col_idx < len(row) else ""
                # Only include non-empty cells
                if cell_value and cell_value.strip():
                    filtered_row.append(cell_value)
            # Only include rows that have at least one non-empty cell
            if filtered_row:
                filtered_data.append(filtered_row)
        
        # Extract headers (first non-empty row)
        headers = []
        data_rows = []
        
        if filtered_data:
            headers = filtered_data[0]
            data_rows = filtered_data[1:] if len(filtered_data) > 1 else []
        
        # Close workbook to free resources
        workbook.close()
        
        return {
            "success": True,
            "sheet_name": sheet_name,
            "headers": headers,
            "data_rows": data_rows,
            "total_rows": total_rows,
            "total_columns": total_columns,
            "filtered_columns": len(non_empty_columns),
            "filtered_rows": len(filtered_data),
            "mode": mode,
            "rows_returned": len(data_rows),
            "error_message": ""
        }
        
    except Exception as e:
        return {
            "success": False,
            "sheet_name": sheet_name,
            "headers": [],
            "data_rows": [],
            "total_rows": 0,
            "total_columns": 0,
            "filtered_columns": 0,
            "filtered_rows": 0,
            "error_message": f"Unexpected error reading sheet content: {str(e)}"
        }


# Tool Definitions for Sheet Identification

GET_SHEET_NAMES_TOOL = ToolDefinition(
    name="GetSheetNames",
    description="Extract all sheet names and basic workbook metadata from an Excel file. Use this tool to get an overview of the workbook structure before analyzing specific sheets.",
    arguments=[
        ToolArgument(
            name="file_path",
            type="string",
            description="Path to the Excel file to analyze"
        )
    ],
    execution_type="activity",
    activity_name="get_sheet_names_activity"
)

READ_SHEET_TOOL = ToolDefinition(
    name="ReadSheet",
    description="Read content from a specific sheet in an Excel file. Returns headers and sample data rows to help identify sheet content and structure.",
    arguments=[
        ToolArgument(
            name="file_path",
            type="string",
            description="Path to the Excel file"
        ),
        ToolArgument(
            name="sheet_name",
            type="string",
            description="Name of the sheet to read"
        ),
        ToolArgument(
            name="mode",
            type="string",
            description="Reading mode: 'preview' (first 15 rows) or 'full' (all rows)"
        )
    ],
    execution_type="activity",
    activity_name="read_sheet_activity"
)

# Tool handlers mapping
SHEET_IDENTIFICATION_TOOL_HANDLERS = {
    "GetSheetNames": get_sheet_names_tool,
    "ReadSheet": read_sheet_tool
}

# List of all sheet identification tool definitions
SHEET_IDENTIFICATION_TOOLS = [
    GET_SHEET_NAMES_TOOL,
    READ_SHEET_TOOL
]