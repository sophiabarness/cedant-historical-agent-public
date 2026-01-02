"""LLM-Primary Catastrophe Data Extraction System.

This module implements LLM-based catastrophe data extraction from Excel sheets.
Sheet identification is handled by the dedicated SheetIdentificationAgent.
"""

import os
import json
from typing import Dict, List, Any
from dataclasses import dataclass, asdict

# Note: NumPy environment variables moved inside functions to comply with Temporal restrictions

from dotenv import load_dotenv
load_dotenv(override=True)

try:
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Catastrophe data extraction from Excel sheets
from models.submission_pack import CatastropheEvent
from .extraction_prompts import load_data_extraction_instructions


@dataclass
class LLMExtractionResult:
    """Result from LLM-guided extraction."""
    success: bool
    events: List[CatastropheEvent]
    extraction_approach: str
    notes: List[str]
    error_message: str = ""


class LLMPrimaryExtractor:
    """LLM-Primary catastrophe data extractor."""
    
    def __init__(self, llm_client=None, custom_extraction_instructions=None):
        """Initialize with LLM client and optional custom prompts.
        
        Args:
            llm_client: LLM client configuration
            custom_extraction_instructions: Custom instructions for data extraction (for GEPA optimization)
        """
        self.llm_client = llm_client or self._get_default_llm_client()
        self.custom_extraction_instructions = custom_extraction_instructions

    
    def _get_default_llm_client(self):
        """Get default LLM client configuration."""
        from litellm import completion
        
        return {
            "model": os.environ.get("LLM_MODEL", "openai/gpt-5"),
            "api_key": os.environ.get("LLM_KEY"),
            "base_url": os.environ.get("LLM_BASE_URL")
        }

    def _call_llm(self, prompt: str, system_prompt: str = None) -> str:
        """Call LLM with the given prompt."""
        if system_prompt is None:
            system_prompt = "You are an expert insurance data analyst specializing in catastrophe loss data extraction from submission packs."
        

        from litellm import completion
        
        messages = [
            {
                "role": "system",
                "content": system_prompt
            },
            {
                "role": "user",
                "content": prompt
            }
        ]
        
        completion_kwargs = {
            "model": self.llm_client["model"],
            "messages": messages,
            "api_key": self.llm_client["api_key"],
            "max_tokens": 16384,  # Sufficient for large sheets with many events (GPT-4o has 128K context)
            "timeout": 120  # 2 minute timeout for LLM calls
        }
        
        # GPT-5 only supports temperature=1, other models can use lower temperatures
        if "gpt-5" not in self.llm_client["model"].lower():
            completion_kwargs["temperature"] = 0.1  # Low temperature for consistent analysis
        
        if self.llm_client.get("base_url"):
            completion_kwargs["base_url"] = self.llm_client["base_url"]
        
        response = completion(**completion_kwargs)
        return response.choices[0].message.content
    
    
    # =====================================================================
    # Data Extraction
    # =====================================================================
    
    def _extract_full_sheet_data(self, worksheet) -> Dict[str, Any]:
        """
        Extract full sheet data for LLM analysis.
        """
        data = {
            "dimensions": f"{worksheet.max_row} rows x {worksheet.max_column} cols",
            "rows": []
        }
        
        # Extract all rows and columns (no artificial limits)
        max_rows = worksheet.max_row or 0
        max_cols = worksheet.max_column or 0
        
        for row in range(1, max_rows + 1):
            row_data = []
            has_content = False
            
            for col in range(1, max_cols + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    # Convert to string, handle various types
                    if isinstance(cell.value, (int, float)):
                        cell_str = str(cell.value)
                    else:
                        cell_str = str(cell.value)[:50]  # Limit length
                    
                    row_data.append(cell_str)
                    has_content = True
                else:
                    row_data.append("")
            
            if has_content:
                data["rows"].append({
                    "row_number": row,
                    "data": row_data
                })
        
        return data
    
    def _parse_extraction_response(self, response: str, default_sheet_name: str) -> List[CatastropheEvent]:
        """Parse LLM extraction response into CatastropheEvent objects."""
        try:
            # Step 1: Handle code blocks if present
            response = response.strip()
            if "```json" in response:
                start_idx = response.find("```json") + 7
                end_idx = response.find("```", start_idx)
                if end_idx != -1:
                    response = response[start_idx:end_idx].strip()
            elif "```" in response:
                start_idx = response.find("```") + 3
                end_idx = response.find("```", start_idx)
                if end_idx != -1:
                    response = response[start_idx:end_idx].strip()
            
            # Step 2: Find JSON array - look for opening [ bracket
            json_start = response.find('[')
            
            if json_start != -1:
                # Find matching closing ] bracket
                json_end = response.rfind(']')
                if json_end != -1:
                    response = response[json_start:json_end+1]
                else:
                    # No closing bracket - try to use from [ onward
                    response = response[json_start:]
            else:
                raise ValueError("No JSON array found in response")
            
            # Parse JSON
            data = json.loads(response)
            
            events = []
            for item in data:
                # Parse loss amount
                loss_gross = item.get("original_loss_gross")
                if loss_gross is not None:
                    try:
                        loss_gross = float(loss_gross)
                    except (ValueError, TypeError):
                        loss_gross = None
                
                # Use source_worksheet from LLM response if available, otherwise use default
                source_worksheet = item.get("source_worksheet") or default_sheet_name
                
                event = CatastropheEvent(
                    loss_year=str(item.get("loss_year")) if item.get("loss_year") else None,
                    loss_description=str(item.get("loss_description")) if item.get("loss_description") else None,
                    original_loss_gross=loss_gross,
                    source_worksheet=source_worksheet,
                    source_row=int(item.get("source_row", 0))
                )
                events.append(event)
            
            return events
            
        except Exception as e:
            print(f"[ERROR] Failed to parse extraction response: {str(e)}")
            return []


    def extract_catastrophe_data(self, file_path: str, sheet_names: List[str], extraction_approach: str, user_instructions: str = None) -> LLMExtractionResult:
        """
        Use LLM to extract catastrophe data with actual calculated values from formulas.
        Concatenates data from all sheets and makes a single LLM call.
        
        Args:
            file_path: Path to Excel file
            sheet_names: List of sheet names to extract from
            extraction_approach: LLM-suggested approach
            user_instructions: Optional user-provided instructions to customize extraction
            
        Returns:
            LLMExtractionResult with extracted events including calculated values
        """
        try:
            workbook_values = load_workbook(file_path, read_only=True, data_only=True)
            
            # Concatenate sheet data from all sheets
            combined_sheet_data = {"sheets": []}
            sheets_processed = []
            
            for sheet_name in sheet_names:
                if sheet_name not in workbook_values.sheetnames:
                    continue
                worksheet_values = workbook_values[sheet_name]
                sheet_data = self._extract_full_sheet_data(worksheet_values)
                combined_sheet_data["sheets"].append({
                    "sheet_name": sheet_name,
                    "data": sheet_data
                })
                sheets_processed.append(sheet_name)
            
            workbook_values.close()
            
            if not sheets_processed:
                return LLMExtractionResult(
                    success=False,
                    events=[],
                    extraction_approach=extraction_approach,
                    notes=[],
                    error_message=f"None of the specified sheets found: {sheet_names}"
                )
            
            # Single LLM call with all sheet data concatenated
            extraction_prompt = self._create_multi_sheet_extraction_prompt(combined_sheet_data, sheets_processed, user_instructions)
            llm_response = self._call_llm(extraction_prompt)
            
            # Parse extraction results
            events = self._parse_extraction_response(llm_response, sheets_processed[0])
            
            result = LLMExtractionResult(
                success=True,
                events=events,
                extraction_approach=extraction_approach,
                notes=[f"LLM extracted {len(events)} events from {len(sheets_processed)} sheets: {', '.join(sheets_processed)}"]
            )
            
            return result
            
        except Exception as e:
            return LLMExtractionResult(
                success=False,
                events=[],
                extraction_approach=extraction_approach,
                notes=[],
                error_message=f"LLM extraction with values failed: {str(e)}"
            )
    
    def _create_multi_sheet_extraction_prompt(self, combined_sheet_data: Dict[str, Any], sheet_names: List[str], user_instructions: str = None) -> str:
        """Create prompt for LLM data extraction from multiple sheets."""
        
        prompt = f"""Extract catastrophe loss events from the following sheets. Data from {len(sheet_names)} sheet(s) is provided below.

Sheets: {', '.join(sheet_names)}

"""
        
        for sheet_info in combined_sheet_data["sheets"]:
            sheet_name = sheet_info["sheet_name"]
            sheet_data = sheet_info["data"]
            
            prompt += f"""
=== SHEET: {sheet_name} ===
Dimensions: {sheet_data['dimensions']}

"""
            for row_info in sheet_data["rows"]:
                non_empty = [cell for cell in row_info["data"] if cell.strip()]
                if non_empty:
                    prompt += f"Row {row_info['row_number']}: {non_empty}\n"
        
        # Load instructions
        instructions = load_data_extraction_instructions(self.custom_extraction_instructions)
        
        prompt += f"""

{instructions}

IMPORTANT: For each event, include "source_worksheet" field indicating which sheet the event came from.
"""
        
        if user_instructions:
            prompt += f"""

ADDITIONAL USER INSTRUCTIONS:
{user_instructions}

Apply these user instructions when extracting the data. They take precedence over default behavior.
"""
        
        return prompt

# Tool interface functions

def llm_extract_catastrophe_data(file_path: str, sheet_names: List[str], extraction_approach: str = "LLM-guided extraction with calculated values", user_instructions: str = None) -> Dict[str, Any]:
    """
    Tool interface for LLM-based catastrophe data extraction with actual calculated values.
    
    Args:
        file_path: Path to submission pack file
        sheet_names: List of sheet names to extract from
        extraction_approach: Extraction approach description
        user_instructions: Optional user-provided instructions to customize extraction
        
    Returns:
        Dictionary with extraction results including calculated loss amounts
    """
    try:
        extractor = LLMPrimaryExtractor()
        result = extractor.extract_catastrophe_data(file_path, sheet_names, extraction_approach, user_instructions)
        
        from dataclasses import asdict
        converted_events = [asdict(event) for event in result.events]
        
        return {
            "success": result.success,
            "events": converted_events,
            "extracted_count": len(result.events),
            "extraction_approach": result.extraction_approach,
            "notes": result.notes,
            "error_message": result.error_message
        }
        
    except Exception as e:
        return {
            "success": False,
            "events": [],
            "error_message": f"LLM extraction with values failed: {str(e)}"
        }